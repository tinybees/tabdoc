#!/usr/bin/env python3
# coding=utf-8

"""
@author: guoyanfeng
@software: PyCharm
@time: 19-2-11 下午6:14
"""
from collections import Counter, MutableMapping, Sequence
from io import BytesIO

import tablib
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from path import Path

__all__ = ("ExcelWriter",)


class ExcelWriter(object):
    """
    excel book writer
    """

    def __init__(self, excel_name, excel_path=None):
        """
            excel book writer
        Args:
            excel_path: excel path
            excel_name: excel 名称
        """
        self.excel_path = excel_path
        self.excel_name = f"{excel_name}.xlsx"
        self.excel_book = tablib.Databook()
        self.merge_cells_index = {}
        self.sheet_names = Counter()  # 多个sheet name的映射，防止名称重复造成错误

    def __enter__(self):
        """

        Args:

        Returns:

        """
        return self

    def __exit__(self, exc_type, exc_val, exc_tb):
        """

        Args:

        Returns:

        """
        self.save()

    # noinspection DuplicatedCode
    @staticmethod
    def _reduce_datetimes(row):
        """Receives a row, converts datetimes to strings."""

        row = list(row)

        for i, val in enumerate(row):
            if hasattr(val, "strftime"):
                row[i] = val.strftime("%Y-%m-%d %H:%M:%S")
            elif hasattr(val, 'isoformat'):
                row[i] = val.isoformat()
        return tuple(row)

    def add_sheet(self, sheet_name, sheet_data: list, merge_cells=None):
        """
        为excel添加工作表
        Args:
            sheet_name: 工作表的名称
            sheet_data: 工作表的数据， 必须是列表中嵌套元祖、列表或者字典（从records查询出来的数据库的数据）
            merge_cells: 要合并的单元格的索引, [(start_row, start_column, end_row, end_column)],最小值从1开始
        Returns:

        """
        sheet_data = sheet_data if sheet_data else [{}]
        #  处理sheet name可能出现重复的情况
        self.sheet_names[sheet_name] += 1
        sheet_name = sheet_name if self.sheet_names[sheet_name] == 1 else f"{sheet_name}{self.sheet_names[sheet_name]}"

        excel_sheet = tablib.Dataset(title=sheet_name)

        for row in sheet_data:
            if not isinstance(row, (MutableMapping, Sequence)):
                raise ValueError("sheet_data值数据类型错误,请检查")

        # 处理list或者tuple个别长度不一致的情况
        first = sheet_data[0]
        if isinstance(first, Sequence):
            for index, row in enumerate(sheet_data[1:], 1):
                diff = len(row) - len(first)
                if abs(diff) > 0:
                    if isinstance(row, list):
                        row.extend(["" for _ in range(diff)])
                    else:
                        sheet_data[index] = (*row, *["" for _ in range(diff)])

        if isinstance(first, MutableMapping):
            excel_sheet.headers = list(first.keys())
            for row in sheet_data:
                row = self._reduce_datetimes(row.values())
                excel_sheet.append(row)
        else:
            excel_sheet.headers = first
            for row in sheet_data[1:]:
                row = self._reduce_datetimes(row)
                excel_sheet.append(row)

        self.excel_book.add_sheet(excel_sheet)
        if merge_cells:
            verify_cells_index = []
            for val in merge_cells:
                verify_cells_index.extend(val)
            if min(verify_cells_index) < 1:
                raise ValueError("Min value is 1")
            self.merge_cells_index[sheet_name] = merge_cells

    # noinspection PyProtectedMember
    def export_book(self, freeze_panes=True):
        """Returns XLSX representation of DataBook."""

        wb = Workbook()
        for sheet in wb.worksheets:
            wb.remove(sheet)
        for i, dset in enumerate(self.excel_book._datasets):
            ws = wb.create_sheet()
            ws.title = dset.title if dset.title else 'Sheet%s' % i
            self.dset_sheet(dset, ws, freeze_panes=freeze_panes)
            # 合并单元格
            if ws.title in self.merge_cells_index:
                for ws_row_col in self.merge_cells_index[ws.title]:
                    ws.merge_cells(start_row=ws_row_col[0], start_column=ws_row_col[1], end_row=ws_row_col[2],
                                   end_column=ws_row_col[3])
                    ws._get_cell(ws_row_col[0], ws_row_col[1]).alignment = Alignment(
                        horizontal="center", vertical="center", wrap_text=True)
        stream = BytesIO()
        wb.save(stream)
        return stream.getvalue()

    # noinspection PyProtectedMember
    @staticmethod
    def dset_sheet(dataset, ws, freeze_panes=True):
        """Completes given worksheet from given Dataset."""
        _package = dataset._package(dicts=False)

        for i, sep in enumerate(dataset._separators):
            _offset = i
            _package.insert((sep[0] + _offset), (sep[1],))

        bold = Font(bold=True)

        for i, row in enumerate(_package):
            row_number = i + 1
            for j, cell_value in enumerate(row):
                col_idx = get_column_letter(j + 1)
                cell = ws['%s%s' % (col_idx, row_number)]
                cell_horizontal, cell_vertical = None, None
                if isinstance(cell_value, dict):
                    cell_color: str = cell_value.get("color", None)
                    # 处理水平居中
                    cell_horizontal: str = cell_value.get("horizontal", None)
                    if cell_horizontal and cell_horizontal not in ("general", "left", "center", "right"):
                        cell_horizontal = "general"  # 默认对其方式

                    # 处理垂直居中
                    cell_vertical: str = cell_value.get("vertical", None)
                    if cell_vertical and cell_vertical not in ("top", "center", "bottom"):
                        cell_vertical = "center"  # 默认对其方式

                    cell_value: str = cell_value.get("value", '')
                    if cell_color:
                        cell.fill = PatternFill("solid", fgColor=cell_color.lstrip("# "))
                cell.alignment = Alignment(wrap_text=True, horizontal=cell_horizontal, vertical=cell_vertical)
                # 增加边框单线，这里是固定的
                thin = Side(border_style="thin", color="000000")
                cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)
                # bold headers
                if (row_number == 1) and dataset.headers:
                    cell.value = str(cell_value)
                    cell.font = bold
                    if freeze_panes:
                        #  Export Freeze only after first Line
                        ws.freeze_panes = 'A2'
                # bold separators
                elif len(row) < dataset.width:
                    cell.value = str(cell_value)
                    cell.font = bold
                # wrap the rest
                else:
                    cell.value = str(cell_value.strip())

    def save(self, ):
        """
        保存工作簿
        Args:
        Returns:

        """
        if self.excel_path is None:
            file_path = self.excel_name
        else:
            file_path = Path(self.excel_path).joinpath(self.excel_name).abspath()

        with open(file_path, "wb") as f:
            f.write(self.export_book())
